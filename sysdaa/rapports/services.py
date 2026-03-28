from __future__ import annotations

import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
from numbers import Number
from urllib.parse import urlencode

from django.core.exceptions import ValidationError
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from articles.models import Article, Categorie
from configurations.models import ConfigurationSysteme
from mouvements_stock.models import MouvementStock
from requisitions.models import LigneRequisition, Requisition
from utilisateurs.models import Direction


MOIS_LABELS = {
    1: "Janvier",
    2: "Février",
    3: "Mars",
    4: "Avril",
    5: "Mai",
    6: "Juin",
    7: "Juillet",
    8: "Août",
    9: "Septembre",
    10: "Octobre",
    11: "Novembre",
    12: "Décembre",
}

STATUT_OK = "OK"
STATUT_ORANGE = "ORANGE"
STATUT_ROUGE = "ROUGE"

TYPE_RAPPORT_LABELS = {
    "stock_global": "État global du stock",
    "stock_global_annuel": "Synthèse annuelle d'activité",
    "categorie_article": "Demandes par catégorie d'article",
    "direction": "Demandes par direction",
    "direction_plus_demandeuse": "Direction la plus demandeuse",
    "direction_moins_demandeuse": "Direction la moins demandeuse",
    "article_plus_demande": "Article le plus demandé",
    "article_moins_demande": "Article le moins demandé",
}


@dataclass(frozen=True)
class ReportRow:
    values: list[str | int | float]


@dataclass(frozen=True)
class SummaryCard:
    label: str
    value: str | int | float
    tone: str = "default"


@dataclass(frozen=True)
class GenericReportData:
    report_type: str
    report_label: str
    period_type: str
    period_label: str
    title: str
    subtitle: str
    annee_fiscale_label: str
    mois: int | None
    mois_label: str
    annee_reelle: int | None
    date_debut: datetime
    date_fin: datetime
    date_export: datetime
    columns: list[str]
    rows: list[ReportRow]
    summary_cards: list[SummaryCard]
    export_filename_base: str
    filters_text: list[str]
    extra_context: dict


@dataclass(frozen=True)
class ReportFilters:
    report_type: str
    period_type: str
    configuration: ConfigurationSysteme
    mois: int | None = None
    categorie: Categorie | None = None
    direction: Direction | None = None
    etat_requisition: str = ""


def _aware_start(d: date) -> datetime:
    tz = timezone.get_current_timezone()
    return timezone.make_aware(datetime.combine(d, time.min), tz)


def _period_for_filters(filters: ReportFilters) -> tuple[datetime, datetime, str, int | None, str]:
    cfg = filters.configuration

    if cfg.annee_debut is None or cfg.annee_fin is None:
        raise ValidationError("Année fiscale invalide.")

    fy_start = date(int(cfg.annee_debut), 10, 1)
    fy_end_exclusive = date(int(cfg.annee_fin), 10, 1)

    if filters.period_type == "ANNUEL":
        return (
            _aware_start(fy_start),
            _aware_start(fy_end_exclusive),
            f"Année fiscale {cfg.code}",
            None,
            "",
        )

    mois = int(filters.mois or 0)
    if mois < 1 or mois > 12:
        raise ValidationError("Mois invalide.")

    annee_reelle = int(cfg.annee_debut) if mois >= 10 else int(cfg.annee_fin)
    start = _aware_start(date(annee_reelle, mois, 1))
    end = (
        _aware_start(date(annee_reelle + 1, 1, 1))
        if mois == 12
        else _aware_start(date(annee_reelle, mois + 1, 1))
    )
    return start, end, f"{MOIS_LABELS[mois]} {annee_reelle}", annee_reelle, MOIS_LABELS[mois]


def _determiner_statut_stock(*, stock_final: int, stock_minimum: int) -> str:
    if stock_final <= 0:
        return STATUT_ROUGE
    if stock_final <= stock_minimum:
        return STATUT_ORANGE
    return STATUT_OK


def _build_filters_text(filters: ReportFilters) -> list[str]:
    parts: list[str] = []

    if filters.categorie is not None:
        parts.append(f"Catégorie : {filters.categorie.libelle}")

    if filters.direction is not None:
        parts.append(f"Direction : {filters.direction.nom}")

    if filters.etat_requisition:
        parts.append(f"État : {filters.etat_requisition}")

    return parts


def _querystring_for_report(filters: ReportFilters) -> str:
    payload = {
        "annee_fiscale": filters.configuration.pk,
        "mois": filters.mois or "",
        "type_rapport": filters.report_type,
        "periode": filters.period_type,
        "etat_requisition": filters.etat_requisition or "",
    }

    if filters.categorie is not None:
        payload["categorie"] = filters.categorie.pk

    if filters.direction is not None:
        payload["direction"] = filters.direction.pk

    return urlencode(payload)


def _base_lignes_queryset(filters: ReportFilters, date_debut: datetime, date_fin: datetime):
    qs = (
        LigneRequisition.objects.select_related(
            "requisition",
            "requisition__soumetteur",
            "requisition__soumetteur__direction_affectee",
            "article",
            "article__categorie",
        )
        .filter(
            requisition__date_preparation__gte=date_debut,
            requisition__date_preparation__lt=date_fin,
        )
    )

    if filters.etat_requisition:
        qs = qs.filter(requisition__etat_requisition=filters.etat_requisition)

    if filters.categorie is not None:
        qs = qs.filter(article__categorie=filters.categorie)

    if filters.direction is not None:
        qs = qs.filter(requisition__soumetteur__direction_affectee=filters.direction)

    return qs


def _base_requisitions_queryset(
    filters: ReportFilters,
    date_debut: datetime,
    date_fin: datetime,
    *,
    date_field: str,
):
    lookup_gte = f"{date_field}__gte"
    lookup_lt = f"{date_field}__lt"

    qs = Requisition.objects.select_related(
        "soumetteur",
        "soumetteur__direction_affectee",
    ).filter(
        **{
            lookup_gte: date_debut,
            lookup_lt: date_fin,
        }
    )

    if filters.etat_requisition:
        qs = qs.filter(etat_requisition=filters.etat_requisition)

    if filters.direction is not None:
        qs = qs.filter(soumetteur__direction_affectee=filters.direction)

    return qs


def _base_mouvements_queryset(filters: ReportFilters, date_debut: datetime, date_fin: datetime):
    qs = (
        MouvementStock.objects.select_related(
            "article",
            "article__categorie",
            "requisition",
            "requisition__soumetteur",
            "requisition__soumetteur__direction_affectee",
        )
        .filter(
            date_mouvement__gte=date_debut,
            date_mouvement__lt=date_fin,
        )
    )

    if filters.categorie is not None:
        qs = qs.filter(article__categorie=filters.categorie)

    if filters.direction is not None:
        qs = qs.filter(requisition__soumetteur__direction_affectee=filters.direction)

    return qs


def _sum_by_article_map(*, qs, type_mouvement: str | None = None) -> dict[int, int]:
    if type_mouvement:
        qs = qs.filter(type_mouvement=type_mouvement)

    data = qs.values("article_id").annotate(total=Sum("quantite_unites"))
    return {
        int(item["article_id"]): int(item["total"] or 0)
        for item in data
        if item["article_id"] is not None
    }


def _safe_pct(numerateur: int, denominateur: int) -> str:
    if denominateur <= 0:
        return "0 %"
    return f"{(numerateur * 100.0 / denominateur):.1f} %"


def _has_requisition_article_filters(filters: ReportFilters) -> bool:
    return filters.direction is not None


def _date_field_for_etat_requisition(etat_requisition: str) -> str:
    etat = (etat_requisition or "").strip()

    if etat == Requisition.ETAT_TRAITEE:
        return "date_livraison"

    if etat in {Requisition.ETAT_VALIDEE, Requisition.ETAT_REJETEE}:
        return "date_approbation"

    return "date_preparation"


def _apply_stock_global_filter_presentation(
    data: GenericReportData,
    filters: ReportFilters,
) -> GenericReportData:
    extra_context = dict(data.extra_context or {})
    filters_text = []

    existing_filters = list(data.filters_text or [])
    categorie_standard = None

    if filters.categorie is not None:
        categorie_standard = f"Catégorie : {filters.categorie.libelle}"
        extra_context["categorie_filtre_label"] = filters.categorie.libelle

        filters_text.append(f"Catégorie filtrée : {filters.categorie.libelle}")

    for item in existing_filters:
        if categorie_standard and item == categorie_standard:
            continue
        if item not in filters_text:
            filters_text.append(item)

    if filters.categorie is not None:
        if data.period_type == "ANNUEL":
            subtitle = (
                "Synthèse annuelle d'activité limitée aux articles de la catégorie sélectionnée, "
                "avec calculs en unités réelles."
            )
        else:
            subtitle = (
                "État global du stock limité aux articles de la catégorie sélectionnée, "
                "avec calculs en unités réelles."
            )
    else:
        subtitle = data.subtitle

    return GenericReportData(
        report_type=data.report_type,
        report_label=data.report_label,
        period_type=data.period_type,
        period_label=data.period_label,
        title=data.title,
        subtitle=subtitle,
        annee_fiscale_label=data.annee_fiscale_label,
        mois=data.mois,
        mois_label=data.mois_label,
        annee_reelle=data.annee_reelle,
        date_debut=data.date_debut,
        date_fin=data.date_fin,
        date_export=data.date_export,
        columns=data.columns,
        rows=data.rows,
        summary_cards=data.summary_cards,
        export_filename_base=data.export_filename_base,
        filters_text=filters_text,
        extra_context=extra_context,
    )


def _generate_stock_global_monthly(
    filters: ReportFilters,
    *,
    date_debut,
    date_fin,
    period_label,
    annee_reelle,
    mois_label,
):
    rows: list[ReportRow] = []

    nb_articles = 0
    nb_orange = 0
    nb_rouge = 0
    nb_ok = 0

    articles = Article.objects.select_related("categorie").order_by("nom", "id")
    if filters.categorie is not None:
        articles = articles.filter(categorie=filters.categorie)

    article_ids = list(articles.values_list("id", flat=True))

    mouvements_periode_qs = MouvementStock.objects.filter(
        article_id__in=article_ids,
        date_mouvement__gte=date_debut,
        date_mouvement__lt=date_fin,
    )
    mouvements_apres_qs = MouvementStock.objects.filter(
        article_id__in=article_ids,
        date_mouvement__gte=date_fin,
    )

    entrees_periode_map = _sum_by_article_map(
        qs=mouvements_periode_qs,
        type_mouvement=MouvementStock.TypeMouvement.ENTREE,
    )
    sorties_periode_map = _sum_by_article_map(
        qs=mouvements_periode_qs,
        type_mouvement=MouvementStock.TypeMouvement.SORTIE,
    )
    entrees_apres_map = _sum_by_article_map(
        qs=mouvements_apres_qs,
        type_mouvement=MouvementStock.TypeMouvement.ENTREE,
    )
    sorties_apres_map = _sum_by_article_map(
        qs=mouvements_apres_qs,
        type_mouvement=MouvementStock.TypeMouvement.SORTIE,
    )

    total_entrees = mouvements_periode_qs.filter(
        type_mouvement=MouvementStock.TypeMouvement.ENTREE
    ).count()
    total_sorties = mouvements_periode_qs.filter(
        type_mouvement=MouvementStock.TypeMouvement.SORTIE
    ).count()

    for article in articles:
        entrees_periode = int(entrees_periode_map.get(article.id, 0))
        sorties_periode = int(sorties_periode_map.get(article.id, 0))
        entrees_apres = int(entrees_apres_map.get(article.id, 0))
        sorties_apres = int(sorties_apres_map.get(article.id, 0))

        stock_final_periode = int(article.stock_actuel or 0) - entrees_apres + sorties_apres
        stock_initial_periode = stock_final_periode - entrees_periode + sorties_periode

        stock_minimum = int(article.stock_minimal or 0)
        statut = _determiner_statut_stock(
            stock_final=stock_final_periode,
            stock_minimum=stock_minimum,
        )

        if statut == STATUT_ROUGE:
            nb_rouge += 1
        elif statut == STATUT_ORANGE:
            nb_orange += 1
        else:
            nb_ok += 1

        rows.append(
            ReportRow(
                values=[
                    article.nom,
                    article.categorie.libelle if article.categorie_id else "-",
                    stock_initial_periode,
                    stock_minimum,
                    entrees_periode,
                    sorties_periode,
                    stock_final_periode,
                    statut,
                ]
            )
        )

        nb_articles += 1

    cards = [
        SummaryCard("Articles", nb_articles),
        SummaryCard("Entrées", total_entrees),
        SummaryCard("Sorties", total_sorties),
        SummaryCard("Alertes vertes", nb_ok, "ok"),
        SummaryCard("Alertes orange", nb_orange, "warning"),
        SummaryCard("Alertes rouges", nb_rouge, "danger"),
    ]

    data = GenericReportData(
        report_type="stock_global",
        report_label=TYPE_RAPPORT_LABELS["stock_global"],
        period_type=filters.period_type,
        period_label=period_label,
        title=f"{TYPE_RAPPORT_LABELS['stock_global']} — {period_label}",
        subtitle="Photographie mensuelle du stock, des entrées, des sorties et des alertes.",
        annee_fiscale_label=filters.configuration.code,
        mois=filters.mois,
        mois_label=mois_label,
        annee_reelle=annee_reelle,
        date_debut=date_debut,
        date_fin=date_fin,
        date_export=timezone.localtime(),
        columns=[
            "Article",
            "Catégorie",
            "Stock initial",
            "Stock min.",
            "Entrées",
            "Sorties",
            "Stock final",
            "État",
        ],
        rows=rows,
        summary_cards=cards,
        export_filename_base=f"rapport_stock_mensuel_{filters.configuration.code.replace('-', '_')}_{filters.mois or ''}",
        filters_text=_build_filters_text(filters),
        extra_context={
            "export_querystring": _querystring_for_report(filters),
            "mode_logique": "stock",
        },
    )
    return _apply_stock_global_filter_presentation(data, filters)


def _generate_stock_global_annual(
    filters: ReportFilters,
    *,
    date_debut,
    date_fin,
    period_label,
    annee_reelle,
    mois_label,
):
    mouvements_qs = _base_mouvements_queryset(filters, date_debut, date_fin)
    lignes_qs = _base_lignes_queryset(filters, date_debut, date_fin)
    req_creees_qs = _base_requisitions_queryset(
        filters,
        date_debut,
        date_fin,
        date_field="date_preparation",
    )
    req_traitees_qs = _base_requisitions_queryset(
        filters,
        date_debut,
        date_fin,
        date_field="date_livraison",
    )

    total_mouvements = mouvements_qs.count()
    total_entrees = int(
        mouvements_qs.filter(
            type_mouvement=MouvementStock.TypeMouvement.ENTREE
        ).aggregate(total=Sum("quantite_unites"))["total"]
        or 0
    )
    total_sorties = int(
        mouvements_qs.filter(
            type_mouvement=MouvementStock.TypeMouvement.SORTIE
        ).aggregate(total=Sum("quantite_unites"))["total"]
        or 0
    )
    total_sorties_manuelles = int(
        mouvements_qs.filter(
            type_mouvement=MouvementStock.TypeMouvement.SORTIE,
            requisition__isnull=True,
        ).aggregate(total=Sum("quantite_unites"))["total"]
        or 0
    )
    total_sorties_requisitions = int(
        mouvements_qs.filter(
            type_mouvement=MouvementStock.TypeMouvement.SORTIE,
            requisition__isnull=False,
        ).aggregate(total=Sum("quantite_unites"))["total"]
        or 0
    )

    nb_requisitions_creees = req_creees_qs.count()
    nb_requisitions_traitees = req_traitees_qs.filter(
        etat_requisition=Requisition.ETAT_TRAITEE
    ).count()
    nb_requisitions_rejetees = req_creees_qs.filter(
        etat_requisition=Requisition.ETAT_REJETEE
    ).count()
    nb_requisitions_en_attente = req_creees_qs.exclude(
        etat_requisition__in=[Requisition.ETAT_TRAITEE, Requisition.ETAT_REJETEE]
    ).count()
    taux_traitement = _safe_pct(nb_requisitions_traitees, nb_requisitions_creees)

    directions_creees = list(
        req_creees_qs.values("soumetteur__direction_affectee__nom")
        .annotate(total=Count("id"))
        .order_by("-total", "soumetteur__direction_affectee__nom")
    )
    directions_traitees = list(
        req_traitees_qs.filter(etat_requisition=Requisition.ETAT_TRAITEE)
        .values("soumetteur__direction_affectee__nom")
        .annotate(total=Count("id"))
        .order_by("-total", "soumetteur__direction_affectee__nom")
    )

    articles_demandes = list(
        lignes_qs.values("article__nom", "article__categorie__libelle")
        .annotate(total=Sum("quantite_demandee_unites"))
        .order_by("-total", "article__nom")[:10]
    )
    articles_livres = list(
        lignes_qs.filter(quantite_livree_unites__gt=0)
        .values("article__nom", "article__categorie__libelle")
        .annotate(total=Sum("quantite_livree_unites"))
        .order_by("-total", "article__nom")[:10]
    )
    articles_sortis = list(
        mouvements_qs.filter(type_mouvement=MouvementStock.TypeMouvement.SORTIE)
        .values("article__nom", "article__categorie__libelle")
        .annotate(total=Sum("quantite_unites"))
        .order_by("-total", "article__nom")[:10]
    )

    rows: list[ReportRow] = []

    rows.extend(
        [
            ReportRow(values=["Vue générale", "Total des mouvements", total_mouvements]),
            ReportRow(values=["Vue générale", "Total des entrées", total_entrees]),
            ReportRow(values=["Vue générale", "Total des sorties", total_sorties]),
            ReportRow(values=["Vue générale", "Sorties manuelles", total_sorties_manuelles]),
            ReportRow(values=["Vue générale", "Sorties liées aux réquisitions", total_sorties_requisitions]),
            ReportRow(values=["Réquisitions", "Réquisitions créées", nb_requisitions_creees]),
            ReportRow(values=["Réquisitions", "Réquisitions traitées", nb_requisitions_traitees]),
            ReportRow(values=["Réquisitions", "Réquisitions rejetées", nb_requisitions_rejetees]),
            ReportRow(values=["Réquisitions", "Réquisitions en cours", nb_requisitions_en_attente]),
            ReportRow(values=["Réquisitions", "Taux simple de traitement", taux_traitement]),
        ]
    )

    if directions_creees:
        for item in directions_creees:
            rows.append(
                ReportRow(
                    values=[
                        "Directions",
                        f"Réquisitions créées — {item['soumetteur__direction_affectee__nom'] or 'Sans direction'}",
                        int(item["total"] or 0),
                    ]
                )
            )

    if directions_traitees:
        for item in directions_traitees:
            rows.append(
                ReportRow(
                    values=[
                        "Directions",
                        f"Réquisitions traitées — {item['soumetteur__direction_affectee__nom'] or 'Sans direction'}",
                        int(item["total"] or 0),
                    ]
                )
            )

    if articles_demandes:
        for item in articles_demandes:
            rows.append(
                ReportRow(
                    values=[
                        "Articles",
                        f"Plus demandés — {item['article__nom']} ({item['article__categorie__libelle'] or 'Sans catégorie'})",
                        int(item["total"] or 0),
                    ]
                )
            )

    if articles_livres:
        for item in articles_livres:
            rows.append(
                ReportRow(
                    values=[
                        "Articles",
                        f"Plus livrés — {item['article__nom']} ({item['article__categorie__libelle'] or 'Sans catégorie'})",
                        int(item["total"] or 0),
                    ]
                )
            )

    if articles_sortis:
        for item in articles_sortis:
            rows.append(
                ReportRow(
                    values=[
                        "Articles",
                        f"Plus sortis — {item['article__nom']} ({item['article__categorie__libelle'] or 'Sans catégorie'})",
                        int(item["total"] or 0),
                    ]
                )
            )

    cards = [
        SummaryCard("Mouvements", total_mouvements),
        SummaryCard("Entrées", total_entrees, "ok"),
        SummaryCard("Sorties", total_sorties, "warning"),
        SummaryCard("Sorties manuelles", total_sorties_manuelles),
        SummaryCard("Sorties via réquisitions", total_sorties_requisitions),
        SummaryCard("Réquisitions créées", nb_requisitions_creees),
        SummaryCard("Réquisitions traitées", nb_requisitions_traitees, "ok"),
        SummaryCard("Réquisitions rejetées", nb_requisitions_rejetees, "danger"),
        SummaryCard("Taux de traitement", taux_traitement),
    ]

    data = GenericReportData(
        report_type="stock_global_annuel",
        report_label=TYPE_RAPPORT_LABELS["stock_global_annuel"],
        period_type=filters.period_type,
        period_label=period_label,
        title=f"{TYPE_RAPPORT_LABELS['stock_global_annuel']} — {period_label}",
        subtitle="Synthèse fiscale des mouvements, réquisitions, directions et articles sur l'année.",
        annee_fiscale_label=filters.configuration.code,
        mois=filters.mois,
        mois_label=mois_label,
        annee_reelle=annee_reelle,
        date_debut=date_debut,
        date_fin=date_fin,
        date_export=timezone.localtime(),
        columns=["Bloc", "Indicateur", "Valeur"],
        rows=rows,
        summary_cards=cards,
        export_filename_base=f"rapport_annuel_activite_{filters.configuration.code.replace('-', '_')}",
        filters_text=_build_filters_text(filters),
        extra_context={
            "export_querystring": _querystring_for_report(filters),
            "mode_logique": "stock_annuel",
        },
    )
    return _apply_stock_global_filter_presentation(data, filters)


def _generate_stock_global(filters: ReportFilters, *, date_debut, date_fin, period_label, annee_reelle, mois_label):
    if filters.period_type == "ANNUEL":
        return _generate_stock_global_annual(
            filters,
            date_debut=date_debut,
            date_fin=date_fin,
            period_label=period_label,
            annee_reelle=annee_reelle,
            mois_label=mois_label,
        )

    return _generate_stock_global_monthly(
        filters,
        date_debut=date_debut,
        date_fin=date_fin,
        period_label=period_label,
        annee_reelle=annee_reelle,
        mois_label=mois_label,
    )


def _generate_requisition_article_breakdown(
    filters: ReportFilters,
    *,
    date_debut,
    date_fin,
    period_label,
    annee_reelle,
    mois_label,
):
    qs = _base_lignes_queryset(filters, date_debut, date_fin)

    data = (
        qs.values(
            "article__nom",
            "article__categorie__libelle",
        )
        .annotate(
            nb_requisitions=Count("requisition", distinct=True),
            quantite_totale_demandee=Sum("quantite_demandee_unites"),
            quantite_totale_livree=Sum("quantite_livree_unites"),
        )
        .order_by("article__nom", "article__categorie__libelle")
    )

    rows = [
        ReportRow(
            values=[
                item["article__nom"] or "-",
                item["article__categorie__libelle"] or "Sans catégorie",
                int(item["nb_requisitions"] or 0),
                int(item["quantite_totale_demandee"] or 0),
                int(item["quantite_totale_livree"] or 0),
            ]
        )
        for item in data
    ]

    cards = [
        SummaryCard("Articles", len(rows)),
        SummaryCard("Réquisitions", sum(int(r.values[2]) for r in rows)),
        SummaryCard("Qté demandée (unités)", sum(int(r.values[3]) for r in rows)),
        SummaryCard("Qté livrée (unités)", sum(int(r.values[4]) for r in rows), "ok"),
    ]

    if filters.direction is not None and filters.categorie is not None:
        subtitle = "Articles de la catégorie sélectionnée demandés par la direction choisie, avec quantités totales demandées et livrées."
    elif filters.direction is not None:
        subtitle = "Articles demandés par la direction choisie, avec quantités totales demandées et livrées."
    elif filters.categorie is not None:
        subtitle = "Articles de la catégorie choisie sur la période, avec quantités totales demandées et livrées."
    else:
        subtitle = "Synthèse des articles demandés sur la période, avec quantités totales demandées et livrées."

    suffix = "annuel" if filters.period_type == "ANNUEL" else "mensuel"

    return GenericReportData(
        report_type="stock_global",
        report_label="Demandes d'articles filtrées",
        period_type=filters.period_type,
        period_label=period_label,
        title=f"Demandes d'articles filtrées — {period_label}",
        subtitle=subtitle,
        annee_fiscale_label=filters.configuration.code,
        mois=filters.mois,
        mois_label=mois_label,
        annee_reelle=annee_reelle,
        date_debut=date_debut,
        date_fin=date_fin,
        date_export=timezone.localtime(),
        columns=["Article", "Catégorie", "Réquisitions", "Qté demandée (unités)", "Qté livrée (unités)"],
        rows=rows,
        summary_cards=cards,
        export_filename_base=f"rapport_articles_filtres_{suffix}_{filters.configuration.code.replace('-', '_')}",
        filters_text=_build_filters_text(filters),
        extra_context={
            "export_querystring": _querystring_for_report(filters),
            "mode_logique": "requisition_article",
        },
    )


def _generate_requisition_state_direction_report(
    filters: ReportFilters,
    *,
    date_debut,
    date_fin,
    period_label,
    annee_reelle,
    mois_label,
):
    date_field = _date_field_for_etat_requisition(filters.etat_requisition)
    req_qs = _base_requisitions_queryset(
        filters,
        date_debut,
        date_fin,
        date_field=date_field,
    )

    lignes_qs = _base_lignes_queryset(filters, date_debut, date_fin)

    data = list(
        req_qs.values("soumetteur__direction_affectee__nom")
        .annotate(
            total_requisitions=Count("id"),
        )
        .order_by("total_requisitions", "soumetteur__direction_affectee__nom")
    )

    lignes_par_direction = {
        (item["requisition__soumetteur__direction_affectee__nom"] or ""): item
        for item in (
            lignes_qs.values("requisition__soumetteur__direction_affectee__nom")
            .annotate(
                quantite_totale_demandee=Sum("quantite_demandee_unites"),
                quantite_totale_livree=Sum("quantite_livree_unites"),
            )
            .order_by("requisition__soumetteur__direction_affectee__nom")
        )
    }

    rows = []
    for item in data:
        direction_nom = item["soumetteur__direction_affectee__nom"] or "Sans direction"
        agregats_lignes = lignes_par_direction.get(
            item["soumetteur__direction_affectee__nom"] or "",
            {},
        )
        rows.append(
            ReportRow(
                values=[
                    direction_nom,
                    int(item["total_requisitions"] or 0),
                    int(agregats_lignes.get("quantite_totale_demandee") or 0),
                    int(agregats_lignes.get("quantite_totale_livree") or 0),
                ]
            )
        )

    etat_label = filters.etat_requisition or "Tous"

    if filters.etat_requisition == Requisition.ETAT_TRAITEE:
        subtitle = "Directions dont les réquisitions ont été traitées sur la période, classées par nombre total croissant."
    elif filters.etat_requisition == Requisition.ETAT_VALIDEE:
        subtitle = "Directions dont les réquisitions ont été validées sur la période, classées par nombre total croissant."
    elif filters.etat_requisition == Requisition.ETAT_REJETEE:
        subtitle = "Directions dont les réquisitions ont été rejetées sur la période, classées par nombre total croissant."
    elif filters.etat_requisition == Requisition.ETAT_EN_ATTENTE_MODIF:
        subtitle = "Directions ayant des réquisitions en attente de modification sur la période, classées par nombre total croissant."
    elif filters.etat_requisition == Requisition.ETAT_EN_ATTENTE:
        subtitle = "Directions ayant des réquisitions en attente sur la période, classées par nombre total croissant."
    else:
        subtitle = "Répartition des réquisitions par direction sur la période, classées par nombre total croissant."

    suffix = "annuel" if filters.period_type == "ANNUEL" else "mensuel"

    return GenericReportData(
        report_type="stock_global",
        report_label=f"Réquisitions {etat_label} par direction",
        period_type=filters.period_type,
        period_label=period_label,
        title=f"Réquisitions {etat_label} par direction — {period_label}",
        subtitle=subtitle,
        annee_fiscale_label=filters.configuration.code,
        mois=filters.mois,
        mois_label=mois_label,
        annee_reelle=annee_reelle,
        date_debut=date_debut,
        date_fin=date_fin,
        date_export=timezone.localtime(),
        columns=["Direction", "Total réquisitions", "Qté demandée (unités)", "Qté livrée (unités)"],
        rows=rows,
        summary_cards=[
            SummaryCard("Directions", len(rows)),
            SummaryCard("Réquisitions", sum(int(r.values[1]) for r in rows)),
            SummaryCard("Qté demandée (unités)", sum(int(r.values[2]) for r in rows)),
            SummaryCard("Qté livrée (unités)", sum(int(r.values[3]) for r in rows), "ok"),
        ],
        export_filename_base=f"rapport_etat_{suffix}_{filters.configuration.code.replace('-', '_')}",
        filters_text=_build_filters_text(filters),
        extra_context={
            "export_querystring": _querystring_for_report(filters),
            "mode_logique": "requisition_etat_direction",
        },
    )


def _generate_categorie_report(
    filters: ReportFilters,
    *,
    date_debut,
    date_fin,
    period_label,
    annee_reelle,
    mois_label,
):
    qs = _base_lignes_queryset(filters, date_debut, date_fin)

    data = (
        qs.values("article__categorie__libelle")
        .annotate(
            quantite_totale=Sum("quantite_demandee_unites"),
            quantite_totale_livree=Sum("quantite_livree_unites"),
            nb_articles=Count("article", distinct=True),
            nb_requisitions=Count("requisition", distinct=True),
        )
        .order_by("article__categorie__libelle")
    )

    rows = [
        ReportRow(
            values=[
                item["article__categorie__libelle"] or "Sans catégorie",
                int(item["nb_articles"] or 0),
                int(item["nb_requisitions"] or 0),
                int(item["quantite_totale"] or 0),
                int(item["quantite_totale_livree"] or 0),
            ]
        )
        for item in data
    ]

    cards = [
        SummaryCard("Catégories listées", len(rows)),
        SummaryCard("Articles distincts", sum(int(r.values[1]) for r in rows)),
        SummaryCard("Réquisitions", sum(int(r.values[2]) for r in rows)),
        SummaryCard("Qté demandée (unités)", sum(int(r.values[3]) for r in rows)),
        SummaryCard("Qté livrée (unités)", sum(int(r.values[4]) for r in rows), "ok"),
    ]

    suffix = "annuel" if filters.period_type == "ANNUEL" else "mensuel"

    return GenericReportData(
        report_type=filters.report_type,
        report_label=TYPE_RAPPORT_LABELS[filters.report_type],
        period_type=filters.period_type,
        period_label=period_label,
        title=f"{TYPE_RAPPORT_LABELS[filters.report_type]} — {period_label}",
        subtitle="Synthèse des demandes regroupées par catégorie d'article.",
        annee_fiscale_label=filters.configuration.code,
        mois=filters.mois,
        mois_label=mois_label,
        annee_reelle=annee_reelle,
        date_debut=date_debut,
        date_fin=date_fin,
        date_export=timezone.localtime(),
        columns=["Catégorie", "Articles distincts", "Réquisitions", "Qté demandée (unités)", "Qté livrée (unités)"],
        rows=rows,
        summary_cards=cards,
        export_filename_base=f"rapport_categories_{suffix}_{filters.configuration.code.replace('-', '_')}",
        filters_text=_build_filters_text(filters),
        extra_context={
            "export_querystring": _querystring_for_report(filters),
            "mode_logique": "categorie",
        },
    )


def _generate_direction_detail_report(
    filters: ReportFilters,
    *,
    date_debut,
    date_fin,
    period_label,
    annee_reelle,
    mois_label,
    mode: str,
):
    qs = _base_lignes_queryset(filters, date_debut, date_fin)

    direction_totaux = list(
        qs.values("requisition__soumetteur__direction_affectee__nom")
        .annotate(
            quantite_totale=Sum("quantite_demandee_unites"),
            quantite_totale_livree=Sum("quantite_livree_unites"),
            nb_articles=Count("article", distinct=True),
            nb_requisitions=Count("requisition", distinct=True),
        )
    )

    if not direction_totaux:
        subtitle = {
            "direction": "Répartition détaillée des demandes par direction.",
            "direction_plus_demandeuse": "Détail de la direction la plus demandeuse.",
            "direction_moins_demandeuse": "Détail de la direction la moins demandeuse.",
        }[mode]
        return GenericReportData(
            report_type=mode,
            report_label=TYPE_RAPPORT_LABELS[mode],
            period_type=filters.period_type,
            period_label=period_label,
            title=f"{TYPE_RAPPORT_LABELS[mode]} — {period_label}",
            subtitle=subtitle,
            annee_fiscale_label=filters.configuration.code,
            mois=filters.mois,
            mois_label=mois_label,
            annee_reelle=annee_reelle,
            date_debut=date_debut,
            date_fin=date_fin,
            date_export=timezone.localtime(),
            columns=["Direction", "Article", "Catégorie", "Réquisitions", "Qté demandée (unités)", "Qté livrée (unités)"],
            rows=[],
            summary_cards=[],
            export_filename_base=f"rapport_{mode}_{'annuel' if filters.period_type == 'ANNUEL' else 'mensuel'}_{filters.configuration.code.replace('-', '_')}",
            filters_text=_build_filters_text(filters),
            extra_context={
                "export_querystring": _querystring_for_report(filters),
                "mode_logique": "direction",
            },
        )

    if mode == "direction_plus_demandeuse":
        direction_totaux.sort(
            key=lambda x: (
                -int(x["quantite_totale"] or 0),
                x["requisition__soumetteur__direction_affectee__nom"] or "",
            )
        )
        direction_cible = direction_totaux[0]
        directions_cibles = [direction_cible["requisition__soumetteur__direction_affectee__nom"]]
    elif mode == "direction_moins_demandeuse":
        direction_totaux.sort(
            key=lambda x: (
                int(x["quantite_totale"] or 0),
                x["requisition__soumetteur__direction_affectee__nom"] or "",
            )
        )
        direction_cible = direction_totaux[0]
        directions_cibles = [direction_cible["requisition__soumetteur__direction_affectee__nom"]]
    else:
        directions_cibles = [
            item["requisition__soumetteur__direction_affectee__nom"]
            for item in sorted(
                direction_totaux,
                key=lambda x: (x["requisition__soumetteur__direction_affectee__nom"] or ""),
            )
        ]

    details_qs = (
        qs.filter(requisition__soumetteur__direction_affectee__nom__in=directions_cibles)
        .values(
            "requisition__soumetteur__direction_affectee__nom",
            "article__nom",
            "article__categorie__libelle",
        )
        .annotate(
            nb_requisitions=Count("requisition", distinct=True),
            quantite_totale=Sum("quantite_demandee_unites"),
            quantite_totale_livree=Sum("quantite_livree_unites"),
        )
        .order_by(
            "requisition__soumetteur__direction_affectee__nom",
            "-quantite_totale",
            "article__nom",
        )
    )

    rows = [
        ReportRow(
            values=[
                item["requisition__soumetteur__direction_affectee__nom"] or "Sans direction",
                item["article__nom"] or "-",
                item["article__categorie__libelle"] or "Sans catégorie",
                int(item["nb_requisitions"] or 0),
                int(item["quantite_totale"] or 0),
                int(item["quantite_totale_livree"] or 0),
            ]
        )
        for item in details_qs
    ]

    if mode == "direction":
        cards = [
            SummaryCard("Directions concernées", len(set(directions_cibles))),
            SummaryCard("Articles distincts", len({(r.values[0], r.values[1]) for r in rows})),
            SummaryCard("Réquisitions", sum(int(r.values[3]) for r in rows)),
            SummaryCard("Qté demandée (unités)", sum(int(r.values[4]) for r in rows)),
            SummaryCard("Qté livrée (unités)", sum(int(r.values[5]) for r in rows), "ok"),
        ]
        subtitle = "Répartition détaillée des demandes par direction et par article."
    else:
        direction_nom = directions_cibles[0] if directions_cibles else "Sans direction"
        total_qte = sum(int(r.values[4]) for r in rows)
        total_qte_livree = sum(int(r.values[5]) for r in rows)
        total_requisitions = sum(int(r.values[3]) for r in rows)
        articles_distincts = len({r.values[1] for r in rows})

        cards = [
            SummaryCard("Direction", direction_nom),
            SummaryCard("Articles distincts", articles_distincts),
            SummaryCard("Réquisitions", total_requisitions),
            SummaryCard("Qté demandée (unités)", total_qte),
            SummaryCard("Qté livrée (unités)", total_qte_livree, "ok"),
        ]
        subtitle = {
            "direction_plus_demandeuse": "Détail de la direction la plus demandeuse avec les articles sollicités.",
            "direction_moins_demandeuse": "Détail de la direction la moins demandeuse avec les articles sollicités.",
        }[mode]

    suffix = "annuel" if filters.period_type == "ANNUEL" else "mensuel"

    return GenericReportData(
        report_type=mode,
        report_label=TYPE_RAPPORT_LABELS[mode],
        period_type=filters.period_type,
        period_label=period_label,
        title=f"{TYPE_RAPPORT_LABELS[mode]} — {period_label}",
        subtitle=subtitle,
        annee_fiscale_label=filters.configuration.code,
        mois=filters.mois,
        mois_label=mois_label,
        annee_reelle=annee_reelle,
        date_debut=date_debut,
        date_fin=date_fin,
        date_export=timezone.localtime(),
        columns=["Direction", "Article", "Catégorie", "Réquisitions", "Qté demandée (unités)", "Qté livrée (unités)"],
        rows=rows,
        summary_cards=cards,
        export_filename_base=f"rapport_{mode}_{suffix}_{filters.configuration.code.replace('-', '_')}",
        filters_text=_build_filters_text(filters),
        extra_context={
            "export_querystring": _querystring_for_report(filters),
            "mode_logique": "direction",
        },
    )


def _generate_article_report(
    filters: ReportFilters,
    *,
    date_debut,
    date_fin,
    period_label,
    annee_reelle,
    mois_label,
    mode: str,
):
    qs = _base_lignes_queryset(filters, date_debut, date_fin)

    data = (
        qs.values("article__nom", "article__categorie__libelle")
        .annotate(
            quantite_totale=Sum("quantite_demandee_unites"),
            quantite_totale_livree=Sum("quantite_livree_unites"),
            nb_demandes=Count("id"),
            nb_requisitions=Count("requisition", distinct=True),
        )
    )

    if mode == "article_plus_demande":
        data = data.order_by("-quantite_totale", "article__nom")[:1]
    else:
        data = data.order_by("quantite_totale", "article__nom")[:1]

    rows = [
        ReportRow(
            values=[
                item["article__nom"],
                item["article__categorie__libelle"] or "Sans catégorie",
                int(item["nb_demandes"] or 0),
                int(item["nb_requisitions"] or 0),
                int(item["quantite_totale"] or 0),
                int(item["quantite_totale_livree"] or 0),
            ]
        )
        for item in data
    ]

    subtitles = {
        "article_plus_demande": "Article le plus demandé sur la période.",
        "article_moins_demande": "Article le moins demandé sur la période.",
    }

    cards = [
        SummaryCard("Articles listés", len(rows)),
        SummaryCard("Nombre de lignes", sum(int(r.values[2]) for r in rows)),
        SummaryCard("Réquisitions", sum(int(r.values[3]) for r in rows)),
        SummaryCard("Qté demandée (unités)", sum(int(r.values[4]) for r in rows)),
        SummaryCard("Qté livrée (unités)", sum(int(r.values[5]) for r in rows), "ok"),
    ]

    suffix = "annuel" if filters.period_type == "ANNUEL" else "mensuel"

    return GenericReportData(
        report_type=mode,
        report_label=TYPE_RAPPORT_LABELS[mode],
        period_type=filters.period_type,
        period_label=period_label,
        title=f"{TYPE_RAPPORT_LABELS[mode]} — {period_label}",
        subtitle=subtitles[mode],
        annee_fiscale_label=filters.configuration.code,
        mois=filters.mois,
        mois_label=mois_label,
        annee_reelle=annee_reelle,
        date_debut=date_debut,
        date_fin=date_fin,
        date_export=timezone.localtime(),
        columns=["Article", "Catégorie", "Nb lignes", "Réquisitions", "Qté demandée (unités)", "Qté livrée (unités)"],
        rows=rows,
        summary_cards=cards,
        export_filename_base=f"rapport_{mode}_{suffix}_{filters.configuration.code.replace('-', '_')}",
        filters_text=_build_filters_text(filters),
        extra_context={
            "export_querystring": _querystring_for_report(filters),
            "mode_logique": "article",
        },
    )


def generer_rapport(filters: ReportFilters) -> GenericReportData:
    date_debut, date_fin, period_label, annee_reelle, mois_label = _period_for_filters(filters)

    if filters.report_type == "stock_global":
        if filters.etat_requisition:
            return _generate_requisition_state_direction_report(
                filters,
                date_debut=date_debut,
                date_fin=date_fin,
                period_label=period_label,
                annee_reelle=annee_reelle,
                mois_label=mois_label,
            )

        if _has_requisition_article_filters(filters):
            return _generate_requisition_article_breakdown(
                filters,
                date_debut=date_debut,
                date_fin=date_fin,
                period_label=period_label,
                annee_reelle=annee_reelle,
                mois_label=mois_label,
            )

        return _generate_stock_global(
            filters,
            date_debut=date_debut,
            date_fin=date_fin,
            period_label=period_label,
            annee_reelle=annee_reelle,
            mois_label=mois_label,
        )

    if filters.report_type == "categorie_article":
        return _generate_categorie_report(
            filters,
            date_debut=date_debut,
            date_fin=date_fin,
            period_label=period_label,
            annee_reelle=annee_reelle,
            mois_label=mois_label,
        )

    if filters.report_type in {"direction", "direction_plus_demandeuse", "direction_moins_demandeuse"}:
        return _generate_direction_detail_report(
            filters,
            date_debut=date_debut,
            date_fin=date_fin,
            period_label=period_label,
            annee_reelle=annee_reelle,
            mois_label=mois_label,
            mode=filters.report_type,
        )

    if filters.report_type in {"article_plus_demande", "article_moins_demande"}:
        return _generate_article_report(
            filters,
            date_debut=date_debut,
            date_fin=date_fin,
            period_label=period_label,
            annee_reelle=annee_reelle,
            mois_label=mois_label,
            mode=filters.report_type,
        )

    raise ValidationError("Type de rapport invalide.")


def generer_rapport_stock_mensuel(*, mois: int):
    cfg = ConfigurationSysteme.objects.filter(est_active=True).order_by("-annee_debut", "-id").first()
    if cfg is None:
        cfg = ConfigurationSysteme.objects.order_by("-annee_debut", "-id").first()
    if cfg is None:
        raise ValidationError("Aucune année fiscale n'est configurée.")

    return generer_rapport(
        ReportFilters(
            report_type="stock_global",
            period_type="MENSUEL",
            configuration=cfg,
            mois=mois,
        )
    )


def _try_find_logo_path() -> str | None:
    candidates = [
        "static/icons/mef.png",
        "sysdaa/static/icons/mef.png",
        "static/images/logo.png",
        "static/img/logo.png",
        "static/core/logo.png",
        "static/logo.png",
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def _xl_fill_for_tone(tone: str) -> PatternFill:
    if tone == "ok":
        return PatternFill("solid", fgColor="EAF7F0")
    if tone == "warning":
        return PatternFill("solid", fgColor="FFF6E8")
    if tone == "danger":
        return PatternFill("solid", fgColor="FFF1F1")
    return PatternFill("solid", fgColor="F8FAFC")


def _xl_color_for_tone(tone: str) -> str:
    if tone == "ok":
        return "179B62"
    if tone == "warning":
        return "D98B00"
    if tone == "danger":
        return "C53B3B"
    return "194993"


def _pdf_column_widths(col_count: int):
    if col_count == 8:
        return [42 * mm, 29 * mm, 16 * mm, 15 * mm, 15 * mm, 15 * mm, 16 * mm, 12 * mm]
    if col_count == 6:
        return [34 * mm, 34 * mm, 22 * mm, 22 * mm, 26 * mm, 26 * mm]
    if col_count == 5:
        return [34 * mm, 42 * mm, 30 * mm, 22 * mm, 34 * mm]
    if col_count == 4:
        return [55 * mm, 30 * mm, 40 * mm, 40 * mm]
    if col_count == 3:
        return [38 * mm, 92 * mm, 32 * mm]
    usable = 170 * mm
    return [usable / max(col_count, 1)] * max(col_count, 1)


def _normalize_total_label(value: str) -> str:
    value = str(value or "").strip().lower()
    value = unicodedata.normalize("NFD", value)
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    return value


def _is_special_monthly_stock_report(data: GenericReportData) -> bool:
    return (
        data.period_type == "MENSUEL"
        and data.report_type == "stock_global"
        and (data.extra_context or {}).get("mode_logique") == "stock"
    )


def _should_render_summary_cards(data: GenericReportData) -> bool:
    return bool(data.summary_cards) and _is_special_monthly_stock_report(data)


def _extract_numeric_for_total(value):
    if isinstance(value, Number):
        return float(value)

    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        if "%" in raw:
            return None
        raw = raw.replace("\xa0", " ").replace(" ", "")
        raw = raw.replace(",", ".")
        if re.fullmatch(r"-?\d+(\.\d+)?", raw):
            try:
                return float(raw)
            except Exception:
                return None
    return None


def _distinct_non_empty(values):
    cleaned = []
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if not text or text == "—":
            continue
        cleaned.append(text)
    return len(set(cleaned))


def _format_total_number(value: float):
    if float(value).is_integer():
        return int(value)
    return round(value, 2)


def _build_total_row(data: GenericReportData):
    if _is_special_monthly_stock_report(data):
        return None

    if not data.rows or not data.columns:
        return None

    headers = list(data.columns)
    normalized_headers = [_normalize_total_label(col) for col in headers]

    if "bloc" in normalized_headers and "indicateur" in normalized_headers:
        return None

    rows = [list(r.values) for r in data.rows]
    total_row = []
    has_metric = False

    first_header = normalized_headers[0]
    first_col_values = [row[0] if len(row) > 0 else None for row in rows]

    if "direction" in first_header:
        count = _distinct_non_empty(first_col_values)
        total_row.append(f"Total ({count} direction(s))" if count else "Total")
        has_metric = has_metric or bool(count)
    elif "article" in first_header:
        count = _distinct_non_empty(first_col_values)
        total_row.append(f"Total ({count} article(s))" if count else "Total")
        has_metric = has_metric or bool(count)
    elif "categorie" in first_header:
        count = _distinct_non_empty(first_col_values)
        total_row.append(f"Total ({count} catégorie(s))" if count else "Total")
        has_metric = has_metric or bool(count)
    else:
        total_row.append("Total")

    for idx in range(1, len(headers)):
        header = normalized_headers[idx]
        col_values = [row[idx] if idx < len(row) else None for row in rows]

        if "etat" in header or "indicateur" in header or "bloc" in header or "taux" in header:
            total_row.append("")
            continue

        if "direction" in header:
            count = _distinct_non_empty(col_values)
            total_row.append(count if count else "")
            has_metric = has_metric or bool(count)
            continue

        if "article" in header:
            count = _distinct_non_empty(col_values)
            total_row.append(count if count else "")
            has_metric = has_metric or bool(count)
            continue

        if "categorie" in header:
            count = _distinct_non_empty(col_values)
            total_row.append(count if count else "")
            has_metric = has_metric or bool(count)
            continue

        numeric_values = []
        for value in col_values:
            numeric = _extract_numeric_for_total(value)
            if numeric is not None:
                numeric_values.append(numeric)

        if numeric_values:
            total_value = sum(numeric_values)
            total_row.append(_format_total_number(total_value))
            has_metric = True
        else:
            total_row.append("")

    return total_row if has_metric else None


def exporter_rapport_excel(*, data: GenericReportData) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"

    fill_dark = PatternFill("solid", fgColor="194993")
    fill_alt = PatternFill("solid", fgColor="F8FAFC")
    fill_total = PatternFill("solid", fgColor="EAF1FB")
    border = Border(
        left=Side(style="thin", color="D7DFEA"),
        right=Side(style="thin", color="D7DFEA"),
        top=Side(style="thin", color="D7DFEA"),
        bottom=Side(style="thin", color="D7DFEA"),
    )

    ws.sheet_view.showGridLines = False
    max_col = max(len(data.columns), 8)

    logo_path = _try_find_logo_path()
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width = 42
            img.height = 42
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
    ws["B1"] = "Ministère de l'Économie et des Finances"
    ws["B1"].font = Font(size=10, bold=True, color="334155")
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=5)
    ws["B2"] = "Direction Générale du Budget / DAA"
    ws["B2"].font = Font(size=10, color="475569")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=max_col)
    ws["A4"] = data.title
    ws["A4"].font = Font(size=18, bold=True, color="194993")
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=max_col)
    ws["A5"] = data.subtitle
    ws["A5"].font = Font(size=10, italic=True, color="6B7280")
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")

    ws["A7"] = f"Mois / Période : {data.period_label}"
    ws["A8"] = f"Année fiscale : {data.annee_fiscale_label}"
    ws["G7"] = f"Exporté le : {timezone.localtime(data.date_export).strftime('%d/%m/%Y %H:%M')}"

    ws["A7"].font = Font(size=10, color="334155")
    ws["A8"].font = Font(size=10, color="334155")
    ws["G7"].font = Font(size=10, color="334155")

    ws["A7"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A8"].alignment = Alignment(horizontal="left", vertical="center")
    ws["G7"].alignment = Alignment(horizontal="right", vertical="center")

    categorie_filtre_label = (data.extra_context or {}).get("categorie_filtre_label")
    table_start_row = 10

    if categorie_filtre_label:
        ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=max_col)
        ws["A9"] = f"Catégorie sélectionnée : {categorie_filtre_label}"
        ws["A9"].font = Font(size=11, bold=True, color="194993")
        ws["A9"].alignment = Alignment(horizontal="left", vertical="center")
        table_start_row = 11

    for idx, title in enumerate(data.columns, start=1):
        cell = ws.cell(row=table_start_row, column=idx, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill_dark
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row_num = table_start_row + 1

    for report_row in data.rows:
        for idx, value in enumerate(report_row.values, start=1):
            display_value = value
            if value in {STATUT_OK, STATUT_ORANGE, STATUT_ROUGE}:
                display_value = "⚑"

            cell = ws.cell(row=row_num, column=idx, value=display_value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if row_num % 2 == 0:
                cell.fill = fill_alt

            if value == STATUT_OK:
                cell.font = Font(bold=True, color="179B62")
            elif value == STATUT_ORANGE:
                cell.font = Font(bold=True, color="D98B00")
            elif value == STATUT_ROUGE:
                cell.font = Font(bold=True, color="C53B3B")
            else:
                cell.font = Font(color="1F2937")

        row_num += 1

    total_row = _build_total_row(data)
    if total_row:
        for idx, value in enumerate(total_row, start=1):
            cell = ws.cell(row=row_num, column=idx, value=value)
            cell.border = border
            cell.fill = fill_total
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, color="194993")
        row_num += 1

    if _should_render_summary_cards(data):
        summary_row = row_num + 2
        start_col = 1
        for idx, card in enumerate(data.summary_cards):
            col = start_col + idx
            if col > 8:
                break
            label_cell = ws.cell(row=summary_row, column=col, value=card.label)
            value_cell = ws.cell(row=summary_row + 1, column=col, value=card.value)

            tone_fill = _xl_fill_for_tone(card.tone)
            tone_color = _xl_color_for_tone(card.tone)

            for c in (label_cell, value_cell):
                c.fill = tone_fill
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            label_cell.font = Font(size=9, bold=True, color="334155")
            value_cell.font = Font(size=11, bold=True, color=tone_color)

    ws.freeze_panes = ws.cell(row=table_start_row + 1, column=1)

    if len(data.columns) == 8:
        widths = {1: 28, 2: 22, 3: 11, 4: 11, 5: 10, 6: 10, 7: 11, 8: 8}
    elif len(data.columns) == 6:
        widths = {1: 24, 2: 24, 3: 14, 4: 16, 5: 16, 6: 16}
    elif len(data.columns) == 5:
        widths = {1: 24, 2: 32, 3: 22, 4: 14, 5: 20}
    elif len(data.columns) == 4:
        widths = {1: 28, 2: 18, 3: 18, 4: 18}
    else:
        widths = {}

    for idx in range(1, max_col + 1):
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = widths.get(idx, 18)

    for r in range(1, row_num + 3):
        ws.row_dimensions[r].height = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{data.export_filename_base}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

def exporter_rapport_pdf(*, data: GenericReportData) -> HttpResponse:
    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()

    style_brand = ParagraphStyle(
        "Brand",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=0,
    )
    style_title = ParagraphStyle(
        "TitlePdf",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#194993"),
        spaceAfter=2,
    )
    style_subtitle = ParagraphStyle(
        "SubtitlePdf",
        parent=styles["BodyText"],
        fontName="Helvetica-Oblique",
        fontSize=9,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#6B7280"),
        spaceAfter=8,
    )
    style_meta_left = ParagraphStyle(
        "MetaLeftPdf",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.7,
        leading=11,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#334155"),
        spaceAfter=1,
    )
    style_meta_right = ParagraphStyle(
        "MetaRightPdf",
        parent=style_meta_left,
        alignment=TA_RIGHT,
    )
    style_filter = ParagraphStyle(
        "FilterPdf",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#194993"),
        spaceAfter=4,
    )
    style_header = ParagraphStyle(
        "HeaderPdf",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.white,
    )
    style_cell_left = ParagraphStyle(
        "CellLeftPdf",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#1F2937"),
    )
    style_cell_center = ParagraphStyle(
        "CellCenterPdf",
        parent=style_cell_left,
        alignment=TA_CENTER,
    )
    style_kpi_label = ParagraphStyle(
        "KpiLabelPdf",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=7.7,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#334155"),
    )
    style_kpi_value = ParagraphStyle(
        "KpiValuePdf",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=13,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#194993"),
    )
    style_total_left = ParagraphStyle(
        "TotalLeftPdf",
        parent=style_cell_left,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#194993"),
    )
    style_total_center = ParagraphStyle(
        "TotalCenterPdf",
        parent=style_cell_center,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#194993"),
    )

    elements = []

    logo_path = _try_find_logo_path()
    logo = ""
    if logo_path:
        try:
            logo = RLImage(logo_path, width=12 * mm, height=12 * mm)
        except Exception:
            logo = ""

    brand_text = Paragraph(
        "Ministère de l'Économie et des Finances<br/>Direction Générale du Budget / DAA",
        style_brand,
    )
    brand_table = Table([[logo, brand_text]], colWidths=[14 * mm, 70 * mm])
    brand_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    elements.append(brand_table)
    elements.append(Spacer(1, 5 * mm))

    elements.append(Paragraph(data.title, style_title))
    elements.append(Paragraph(data.subtitle, style_subtitle))

    meta_left_rows = [
        [Paragraph(f"Mois / Période : {data.period_label}", style_meta_left)],
        [Paragraph(f"Année fiscale : {data.annee_fiscale_label}", style_meta_left)],
    ]

    categorie_filtre_label = (data.extra_context or {}).get("categorie_filtre_label")
    if categorie_filtre_label:
        meta_left_rows.append(
            [Paragraph(f"Catégorie sélectionnée : {categorie_filtre_label}", style_filter)]
        )

    meta_left = Table(
        meta_left_rows,
        colWidths=[78 * mm],
    )
    meta_left.setStyle(
        TableStyle(
            [
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ]
        )
    )

    meta_right = Table(
        [[Paragraph(f"Exporté le : {timezone.localtime(data.date_export).strftime('%d/%m/%Y %H:%M')}", style_meta_right)]],
        colWidths=[72 * mm],
    )
    meta_right.setStyle(
        TableStyle(
            [
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ]
        )
    )

    meta_table = Table([[meta_left, meta_right]], colWidths=[82 * mm, 82 * mm])
    meta_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    elements.append(meta_table)
    elements.append(Spacer(1, 4 * mm))

    table_data = [[Paragraph(str(col), style_header) for col in data.columns]]

    for row in data.rows:
        pdf_row = []
        for idx, value in enumerate(row.values):
            style = style_cell_left if idx == 0 else style_cell_center
            if value == STATUT_OK:
                pdf_row.append(Paragraph('<font color="#179B62">⚑</font>', style_cell_center))
            elif value == STATUT_ORANGE:
                pdf_row.append(Paragraph('<font color="#D98B00">⚑</font>', style_cell_center))
            elif value == STATUT_ROUGE:
                pdf_row.append(Paragraph('<font color="#C53B3B">⚑</font>', style_cell_center))
            else:
                pdf_row.append(Paragraph(str(value), style))
        table_data.append(pdf_row)

    total_row = _build_total_row(data)
    has_total_row = False
    if total_row:
        has_total_row = True
        pdf_total_row = []
        for idx, value in enumerate(total_row):
            style = style_total_left if idx == 0 else style_total_center
            pdf_total_row.append(Paragraph(str(value), style))
        table_data.append(pdf_total_row)

    report_table = Table(
        table_data,
        colWidths=_pdf_column_widths(len(data.columns)),
        repeatRows=1,
    )

    table_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#194993")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D7DFEA")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    if has_total_row:
        last_row = len(table_data) - 1
        table_styles.extend(
            [
                ("BACKGROUND", (0, last_row), (-1, last_row), colors.HexColor("#EAF1FB")),
                ("FONTNAME", (0, last_row), (-1, last_row), "Helvetica-Bold"),
            ]
        )

    report_table.setStyle(TableStyle(table_styles))
    elements.append(report_table)
    elements.append(Spacer(1, 4 * mm))

    if _should_render_summary_cards(data):
        kpi_cells = []
        for card in data.summary_cards:
            tone_bg = "#F8FAFC"
            tone_value = "#194993"
            if card.tone == "ok":
                tone_bg = "#EAF7F0"
                tone_value = "#179B62"
            elif card.tone == "warning":
                tone_bg = "#FFF6E8"
                tone_value = "#D98B00"
            elif card.tone == "danger":
                tone_bg = "#FFF1F1"
                tone_value = "#C53B3B"

            label = Paragraph(str(card.label), style_kpi_label)
            value = Paragraph(f'<font color="{tone_value}">{card.value}</font>', style_kpi_value)
            cell_table = Table([[label], [value]], colWidths=[28 * mm])
            cell_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(tone_bg)),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D7DFEA")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]
                )
            )
            kpi_cells.append(cell_table)

        kpi_table = Table([kpi_cells], colWidths=[28 * mm] * len(kpi_cells))
        kpi_table.setStyle(
            TableStyle(
                [
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        elements.append(kpi_table)

    doc.build(elements)

    pdf = output.getvalue()
    output.close()

    filename = f"{data.export_filename_base}.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response