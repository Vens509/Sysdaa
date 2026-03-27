from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Set, Tuple

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from articles.models import Article, Categorie


class Command(BaseCommand):
    help = (
        "Importe les articles depuis un fichier Excel en ne gardant "
        "que le nom de l'article et sa catégorie."
    )

    SHEET_CATEGORY_MAPPING: Dict[str, str] = {
        "Produits de nettoyage": "Produits de nettoyage",
        "Articles divers": "Articles divers",
        "Fournitures de bur.": "Fournitures de bureau",
        "Toner et Drum": "Materiels Informatiques et electroniques",
    }

    CATEGORY_ROW_MARKERS: Dict[str, str] = {
        "Produits de nettoyage": "Produits de nettoyage",
        "Articles divers": "Articles divers",
        "Fournitures de bur.": "Fournitures de bureau",
        "Toner et Drum": "Toner ;Drum & Clé USB",
    }

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Chemin complet du fichier Excel à importer.",
        )
        parser.add_argument(
            "--user-email",
            dest="user_email",
            required=True,
            help="Email de l'utilisateur à enregistrer comme createur des articles.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Affiche ce qui serait importé sans écrire dans la base.",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"]).expanduser().resolve()
        user_email = (options["user_email"] or "").strip()
        dry_run = options["dry_run"]

        if not excel_path.exists():
            raise CommandError(f"Fichier introuvable : {excel_path}")

        if excel_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            raise CommandError("Le fichier doit être un classeur Excel valide.")

        User = get_user_model()
        utilisateur = User.objects.filter(email__iexact=user_email).first()
        if not utilisateur:
            raise CommandError(
                f"Aucun utilisateur trouvé avec l'email : {user_email}"
            )

        try:
            workbook = load_workbook(excel_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"Impossible de lire le fichier Excel : {exc}") from exc

        feuilles_manquantes = [
            sheet_name
            for sheet_name in self.SHEET_CATEGORY_MAPPING
            if sheet_name not in workbook.sheetnames
        ]
        if feuilles_manquantes:
            raise CommandError(
                "Feuilles manquantes dans le fichier Excel : "
                + ", ".join(feuilles_manquantes)
            )

        categories_cache: Dict[str, Categorie] = {}
        articles_a_creer: List[Article] = []
        noms_deja_vus_dans_excel: Set[str] = set()

        total_lus = 0
        total_ignores_excel = 0
        total_ignores_db = 0

        # Précharger les noms déjà existants en base (insensibles à la casse)
        noms_existants_db = {
            nom.casefold()
            for nom in Article.objects.values_list("nom", flat=True)
        }

        for sheet_name, categorie_libelle in self.SHEET_CATEGORY_MAPPING.items():
            worksheet = workbook[sheet_name]
            nom_categorie_attendu = self.CATEGORY_ROW_MARKERS[sheet_name]

            categorie = categories_cache.get(categorie_libelle)
            if categorie is None:
                categorie, _ = Categorie.objects.get_or_create(libelle=categorie_libelle)
                categories_cache[categorie_libelle] = categorie

            noms_extraits = self._extract_article_names_from_sheet(
                worksheet=worksheet,
                expected_category_row=nom_categorie_attendu,
            )

            self.stdout.write("")
            self.stdout.write(self.style.NOTICE(f"Feuille : {sheet_name}"))
            self.stdout.write(self.style.NOTICE(f"Catégorie : {categorie_libelle}"))

            for nom_article in noms_extraits:
                total_lus += 1
                nom_normalise = self._normalize_key(nom_article)

                if nom_normalise in noms_deja_vus_dans_excel:
                    total_ignores_excel += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"Doublon ignoré dans Excel : {nom_article}"
                        )
                    )
                    continue

                if nom_normalise in noms_existants_db:
                    total_ignores_db += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"Déjà existant en base, ignoré : {nom_article}"
                        )
                    )
                    noms_deja_vus_dans_excel.add(nom_normalise)
                    continue

                article = Article(
                    nom=nom_article,
                    unite="Unité",
                    unite_base="Unité",
                    quantite_par_conditionnement=1,
                    stock_initial=0,
                    stock_actuel=0,
                    stock_minimal=0,
                    categorie=categorie,
                    utilisateur_enregistreur=utilisateur,
                )

                articles_a_creer.append(article)
                noms_deja_vus_dans_excel.add(nom_normalise)

        self.stdout.write("")
        self.stdout.write(self.style.NOTICE("Résumé de l'import"))
        self.stdout.write(f"Articles lus : {total_lus}")
        self.stdout.write(f"Doublons ignorés dans Excel : {total_ignores_excel}")
        self.stdout.write(f"Déjà présents en base : {total_ignores_db}")
        self.stdout.write(f"Nouveaux articles à créer : {len(articles_a_creer)}")

        if dry_run:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Mode dry-run activé : aucune insertion effectuée."))
            return

        if not articles_a_creer:
            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("Aucun nouvel article à insérer."))
            return

        with transaction.atomic():
            Article.objects.bulk_create(articles_a_creer, batch_size=500)

        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                f"Import terminé avec succès : {len(articles_a_creer)} article(s) inséré(s)."
            )
        )

    def _extract_article_names_from_sheet(self, worksheet, expected_category_row: str) -> List[str]:
        """
        Extrait les noms d'articles depuis la colonne B.

        Logique utilisée :
        1. Trouver la ligne contenant 'Liste des articles'
        2. Vérifier que la ligne suivante contient bien le nom de la catégorie attendu
        3. Lire les lignes suivantes dans la colonne B jusqu'à rencontrer
           une ligne vide ou un nouveau bloc logique
        """
        article_names: List[str] = []

        row_liste_articles = self._find_row_in_column_b(worksheet, "Liste des articles")
        if row_liste_articles is None:
            raise CommandError(
                f"Impossible de trouver 'Liste des articles' dans la feuille '{worksheet.title}'."
            )

        row_categorie = row_liste_articles + 1
        valeur_categorie = self._clean_cell_value(worksheet.cell(row=row_categorie, column=2).value)

        if not valeur_categorie:
            raise CommandError(
                f"Impossible de trouver la ligne de catégorie dans la feuille '{worksheet.title}'."
            )

        # Contrôle souple, pour éviter de casser si le libellé varie légèrement.
        if self._normalize_key(expected_category_row) not in self._normalize_key(valeur_categorie):
            raise CommandError(
                f"La catégorie détectée dans la feuille '{worksheet.title}' "
                f"ne correspond pas à celle attendue. "
                f"Attendu proche de : '{expected_category_row}' | Trouvé : '{valeur_categorie}'"
            )

        row_start = row_categorie + 1

        for row_idx in range(row_start, worksheet.max_row + 1):
            raw_value = worksheet.cell(row=row_idx, column=2).value
            value = self._clean_cell_value(raw_value)

            if not value:
                # On arrête seulement après avoir commencé à lire des articles.
                if article_names:
                    break
                continue

            normalized = self._normalize_key(value)

            if normalized in {
                "liste des articles",
                "articles divers",
                "produits de nettoyage",
                "fournitures de bureau",
                "toner et drum",
                "toner ;drum & clé usb",
            }:
                continue

            if normalized.startswith("pour le mois"):
                break

            if normalized.startswith("exercice:"):
                break

            if normalized.startswith("service des ressources matérielles"):
                break

            if normalized.startswith("services des ressources matérielles"):
                break

            article_names.append(value)

        return article_names

    @staticmethod
    def _find_row_in_column_b(worksheet, searched_text: str) -> int | None:
        searched = Command._normalize_key(searched_text)
        for row_idx in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=2).value
            cleaned = Command._clean_cell_value(value)
            if cleaned and Command._normalize_key(cleaned) == searched:
                return row_idx
        return None

    @staticmethod
    def _clean_cell_value(value) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        return " ".join(text.split())

    @staticmethod
    def _normalize_key(value: str) -> str:
        return " ".join((value or "").strip().split()).casefold()